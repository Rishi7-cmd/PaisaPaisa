import streamlit as st
import pandas as pd
import base64
import io
import time
from openpyxl import Workbook

# 🌃 Diwali-inspired dark room theme
st.markdown("""
    <style>
        body {
            background-color: #0e0e0e;
            background-image: radial-gradient(#ffcc00 0.5px, #0e0e0e 1px);
            background-size: 20px 20px;
        }
        .stApp {
            background-color: #0e0e0e;
            color: #f1f1f1;
        }
        .title-text {
            font-size: 48px;
            font-weight: 800;
            color: #ffd700;
            text-shadow: 0 0 5px #ff9933, 0 0 10px #ff6600, 0 0 15px #cc3300;
            text-align: center;
            margin-top: 30px;
        }
        .subtext {
            text-align: center;
            font-size: 18px;
            color: #dddddd;
            margin-bottom: 30px;
        }
        .css-1aumxhk {
            background-color: #111 !important;
        }
        .stButton>button {
            background-color: #ff9900 !important;
            color: white !important;
        }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="title-text">🪔 PaisaPaisa Layered Transaction<br>Flowchart</div>', unsafe_allow_html=True)
st.markdown('<div class="subtext">Upload a transaction Excel file and get a processed flowchart-style Excel as output.</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

def process_excel(file):
    df = pd.read_excel(file)

    # Filter > ₹50,000 and withdrawals > ₹1L
    df = df[df['Amount'] > 50000]

    wb = Workbook()
    ws = wb.active
    ws.title = "Flowchart"

    headers = ["Victim", "Layer 1", "Layer 2", "Withdrawal", "Amount"]
    ws.append(headers)

    for i in range(len(df)):
        row = df.iloc[i]
        layer2 = row['Layer 2 Account'] if pd.notna(row['Layer 2 Account']) else ""
        withdrawal = "Yes" if row['Withdrawal'] == "Yes" else "No"
        amt = row['Amount']
        out_row = [row['Victim Account'], row['Layer 1 Account'], layer2, withdrawal, amt]

        ws.append(out_row)

        if amt > 100000:
            for col in range(1, 6):
                ws.cell(row=ws.max_row, column=col).fill = openpyxl.styles.PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

if uploaded_file:
    with st.spinner("Generating your layered flowchart... 💡"):
        result = process_excel(uploaded_file)
        st.success("Done! Download below:")
        st.download_button("📥 Download Flowchart Excel", result, file_name="flowchart_output.xlsx")
