import pandas as pd
import re
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def process_excel(uploaded_file):
    df = pd.read_excel(uploaded_file)
    
    # Preprocessing
    df['Processed_Amount'] = df['Amount'] * 1.18  # GST example
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Layer'] = df['Account_Type'].map({'Victim': 0, 'Layer1': 1, 'Layer2': 2, 'Layer3': 3})

    # Withdrawal filter
    df['Withdrawal_Flag'] = df['Description'].str.contains('withdraw|atm|cash', case=False, na=False)
    df['High_Withdrawal'] = df.apply(lambda x: x['Amount'] if x['Withdrawal_Flag'] and x['Amount'] > 100000 else None, axis=1)

    # Filter: Amount > 50,000 and only till Layer 2
    filtered_df = df[(df['Amount'] > 50000) & (df['Layer'] <= 2)]

    # Retain only Layer2 rows that have withdrawals
    filtered_df = filtered_df[~((filtered_df['Layer'] == 2) & (~filtered_df['Withdrawal_Flag']))]

    # Prepare flowchart layout
    flow = {}
    for _, row in filtered_df.iterrows():
        src = row['From_Account']
        dst = row['To_Account']
        amt = f"₹{int(row['Amount'])}"
        flow.setdefault(src, []).append(f"{dst} ({amt})")

    # Generate flow-style table
    max_depth = max([len(v) for v in flow.values()] + [1])
    table = []
    for i in range(max_depth):
        row = []
        for k in flow.keys():
            try:
                row.append(flow[k][i])
            except:
                row.append("")
        table.append(row)

    flow_df = pd.DataFrame(table, columns=flow.keys())

    # Excel export
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"{os.path.splitext(uploaded_file.name)[0]}_output.xlsx"
    output_path = os.path.join("/mount/app", output_filename)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        flow_df.to_excel(writer, sheet_name='FlowChart', index=False)
        filtered_df.to_excel(writer, sheet_name='Filtered_Data', index=False)

        workbook = writer.book
        flow_sheet = workbook['FlowChart']
        data_sheet = workbook['Filtered_Data']

        # Style headers
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for cell in flow_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for cell in data_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Highlight large withdrawals
        for row in range(2, data_sheet.max_row + 1):
            amt_cell = data_sheet[f"L{row}"]
            if amt_cell.value and amt_cell.value > 100000:
                amt_cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    return output_path
